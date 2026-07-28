#!/usr/bin/env python3
"""
process_cards.py

Verarbeitet die PlayPointy Kartendaten aus der Excel-Datei und ordnet die
Bilddateien aus dem Packs/-Ordner den einzelnen Karten zu, per Fuzzy-Matching
(difflib.SequenceMatcher) zwischen Bilddateinamen und Excel-Kartentexten.

Modus 1 (Standard / DRY RUN):
    python process_cards.py
    -> Liest Excel-Tabelle + Ordnerstruktur ein und schreibt NUR
       "mapping_check.txt" zur manuellen Kontrolle.
       Es werden KEINE Dateien kopiert oder veraendert.

Modus 2 (EXECUTE):
    python process_cards.py --execute
    -> Kopiert die zugeordneten Bilder nach public/cards/card_001.webp bis
       card_150.webp und erstellt public/cards.json sowie public/packs.json.

Hinweis zu den Zeilenbereichen:
    Die Start-/End-Zeilen jedes Packs werden NICHT hart einprogrammiert,
    sondern automatisch erkannt (Pack-Header-Zelle suchen, dann Zeilen lesen
    bis zur ersten leeren Zelle). Das macht das Skript robust gegenueber
    Verschiebungen in der Excel-Datei. Die tatsaechlich erkannten Bereiche
    werden am Anfang von mapping_check.txt protokolliert.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import column_index_from_string

# --------------------------------------------------------------------------
# Konfiguration
# --------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "PlayPointy (1).xlsx"
PACKS_DIR = BASE_DIR / "Packs"
SHEET_NAME = "Packs"
OUTPUT_DIR = BASE_DIR / "public"
CARDS_OUT_DIR = OUTPUT_DIR / "cards"
MAPPING_FILE = BASE_DIR / "mapping_check.txt"
CARDS_JSON = OUTPUT_DIR / "cards.json"
PACKS_JSON = OUTPUT_DIR / "packs.json"

IMAGE_EXTENSIONS = {".webp", ".png", ".jpg", ".jpeg"}

# Reihenfolge bestimmt die fortlaufende card_XXX Nummerierung (001-150).
PACK_DEFS = [
    {"name": "Starter Chaos", "card_col": "B", "hex_col": "C"},
    {"name": "Dark & Evil", "card_col": "F", "hex_col": "G"},
    {"name": "Roast Friends", "card_col": "I", "hex_col": "J"},
    {"name": "Toxic Love", "card_col": "B", "hex_col": "C"},
    {"name": "Unhinged Nights", "card_col": "F", "hex_col": "G"},
]

WARN_THRESHOLD = 70.0  # % - darunter wird eine Warnung in mapping_check.txt ausgegeben


@dataclass
class ExcelCard:
    global_id: int
    pack: str
    row: int
    text: str
    hex_raw: Optional[object]
    hex_norm: Optional[str]
    assigned_image: Optional["ImageFile"] = None
    match_score: float = 0.0


@dataclass
class ImageFile:
    path: Path
    used: bool = False


# --------------------------------------------------------------------------
# Hilfsfunktionen
# --------------------------------------------------------------------------

def normalize_hex(value) -> Optional[str]:
    """Normalisiert Excel Hex-Werte zu '#rrggbb'.

    Excel wandelt rein numerische Hex-Codes (z.B. '9c9c00') teilweise
    automatisch in Zahlen um und verschluckt dabei fuehrende Nullen
    (aus '007165' wird die Zahl 7165). Das wird hier wieder aufgefuellt.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            ival = int(round(value))
        except (ValueError, OverflowError):
            return None
        s = str(ival)
    else:
        s = str(value).strip()
        if s.startswith("#"):
            s = s[1:]
    s = s.strip()
    if not s:
        return None
    if len(s) < 6:
        s = s.zfill(6)
    return "#" + s.lower()


_PAREN_RE = re.compile(r"\([^)]*\)")
_PUNCT_RE = re.compile(r"[?!.,\"'\u2019\u2018\u201c\u201d]")
_WS_RE = re.compile(r"\s+")


def clean_text(text: str) -> str:
    """Vereinheitlicht Kartentext/Dateinamen fuer den Fuzzy-Vergleich
    (entfernt z.B. Klammerzusaetze wie '(Unhinged Nights)', Satzzeichen,
    Unter-/Bindestriche, Mehrfach-Leerzeichen, Gross-/Kleinschreibung)."""
    t = _PAREN_RE.sub(" ", text)
    t = t.replace("_", " ").replace("-", " ")
    t = _PUNCT_RE.sub("", t)
    t = _WS_RE.sub(" ", t)
    return t.strip().lower()


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio() * 100.0


# --------------------------------------------------------------------------
# Excel einlesen
# --------------------------------------------------------------------------

def find_header_row(ws, col_idx: int, pack_name: str) -> Optional[int]:
    target = pack_name.strip().lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is not None and str(v).strip().lower() == target:
            return r
    return None


def extract_pack_cards(ws, pack_name: str, card_col: str, hex_col: str, start_id: int):
    """Liest alle Kartenzeilen eines Packs aus. Start-/Endzeile werden
    automatisch erkannt: Header-Zelle mit dem Pack-Namen suchen, danach
    leere Zeilen ueberspringen, dann Zeilen lesen bis zur ersten Luecke."""
    card_col_idx = column_index_from_string(card_col)
    hex_col_idx = column_index_from_string(hex_col)

    header_row = find_header_row(ws, card_col_idx, pack_name)
    if header_row is None:
        raise ValueError(f"Konnte Header '{pack_name}' in Spalte {card_col} nicht finden.")

    r = header_row + 1
    max_row = ws.max_row
    while r <= max_row:
        v = ws.cell(row=r, column=card_col_idx).value
        if v is not None and str(v).strip():
            break
        r += 1
    start_row = r

    cards = []
    gid = start_id
    while r <= max_row:
        v = ws.cell(row=r, column=card_col_idx).value
        if v is None or not str(v).strip():
            break
        hex_raw = ws.cell(row=r, column=hex_col_idx).value
        cards.append(
            ExcelCard(
                global_id=gid,
                pack=pack_name,
                row=r,
                text=str(v).strip(),
                hex_raw=hex_raw,
                hex_norm=normalize_hex(hex_raw),
            )
        )
        gid += 1
        r += 1
    end_row = r - 1
    return cards, start_row, end_row


def load_all_cards(excel_path: Path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Tabellenblatt '{SHEET_NAME}' nicht gefunden. Vorhanden: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    all_cards: list[ExcelCard] = []
    pack_ranges = []
    next_id = 1
    for pdef in PACK_DEFS:
        cards, start_row, end_row = extract_pack_cards(
            ws, pdef["name"], pdef["card_col"], pdef["hex_col"], next_id
        )
        all_cards.extend(cards)
        pack_ranges.append(
            {
                "name": pdef["name"],
                "card_col": pdef["card_col"],
                "hex_col": pdef["hex_col"],
                "start_row": start_row,
                "end_row": end_row,
                "count": len(cards),
            }
        )
        next_id += len(cards)
    return all_cards, pack_ranges


# --------------------------------------------------------------------------
# Bilder einlesen
# --------------------------------------------------------------------------

def collect_images(packs_dir: Path) -> list[ImageFile]:
    images = []
    if not packs_dir.exists():
        return images
    for p in sorted(packs_dir.rglob("*")):
        if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS:
            images.append(ImageFile(path=p))
    return images


# --------------------------------------------------------------------------
# Fuzzy-Matching (difflib.SequenceMatcher), global & eindeutig (1:1)
# --------------------------------------------------------------------------

def match_cards_to_images(cards: list[ExcelCard], images: list[ImageFile]):
    card_clean = {c.global_id: clean_text(c.text) for c in cards}
    image_clean = {id(img): clean_text(img.path.stem) for img in images}

    pairs = []
    for c in cards:
        ct = card_clean[c.global_id]
        for img in images:
            it = image_clean[id(img)]
            pairs.append((similarity(ct, it), c, img))

    # Groesste Aehnlichkeit zuerst zuweisen (greedy), jede Karte/jedes Bild
    # wird dabei nur genau einmal verwendet.
    pairs.sort(key=lambda x: x[0], reverse=True)

    assigned_cards: set[int] = set()
    assigned_images: set[int] = set()
    for score, card, img in pairs:
        if card.global_id in assigned_cards or id(img) in assigned_images:
            continue
        card.assigned_image = img
        card.match_score = score
        img.used = True
        assigned_cards.add(card.global_id)
        assigned_images.add(id(img))
        if len(assigned_cards) == len(cards) or len(assigned_images) == len(images):
            break


# --------------------------------------------------------------------------
# Modus 1: DRY RUN -> mapping_check.txt
# --------------------------------------------------------------------------

def write_mapping_check(cards, images, pack_ranges, mapping_file: Path):
    lines = []
    lines.append("=" * 100)
    lines.append("PLAYPOINTY - KARTEN/BILDER MAPPING CHECK (DRY RUN)")
    lines.append("=" * 100)
    lines.append("")
    lines.append(f"Excel-Datei  : {EXCEL_PATH.name}")
    lines.append(f"Bilder-Ordner: {PACKS_DIR}")
    lines.append(f"Gefundene Bilder gesamt: {len(images)}")
    lines.append(f"Gefundene Karten gesamt: {len(cards)}")
    lines.append("")
    lines.append("Erkannte Pack-Bereiche (automatisch anhand Header + erster Luecke erkannt):")
    for pr in pack_ranges:
        lines.append(
            f"  - {pr['name']:<18} Spalte {pr['card_col']}/{pr['hex_col']}  "
            f"Zeilen {pr['start_row']}-{pr['end_row']}  ({pr['count']} Karten)"
        )
    lines.append("")
    lines.append("-" * 100)
    lines.append("ZUORDNUNGEN")
    lines.append("-" * 100)

    warnings = []
    no_match = []
    current_pack = None
    for c in cards:
        if c.pack != current_pack:
            current_pack = c.pack
            lines.append("")
            lines.append(f"### Pack: {current_pack} ###")
        target_id = f"card_{c.global_id:03d}.webp"
        hex_display = c.hex_norm if c.hex_norm else "KEIN HEX"
        if c.assigned_image is not None:
            orig_path = c.assigned_image.path.relative_to(BASE_DIR).as_posix()
            match_str = f"{c.match_score:.1f}%"
            if c.match_score < WARN_THRESHOLD:
                warnings.append(
                    f"  Zeile {c.row} ({c.pack}): '{c.text}' -> {orig_path} nur {match_str} Uebereinstimmung"
                )
        else:
            orig_path = "!! KEIN BILD GEFUNDEN !!"
            match_str = "0.0%"
            no_match.append(f"  Zeile {c.row} ({c.pack}): '{c.text}'")

        lines.append(
            f"[{orig_path}] -> [Ziel-ID: {target_id}] | Match: {match_str} | "
            f"Hex: {hex_display} | Text aus Excel: {c.text}"
        )

    unused_images = [img for img in images if not img.used]

    lines.append("")
    lines.append("-" * 100)
    lines.append("WARNUNGEN")
    lines.append("-" * 100)
    if not warnings and not no_match:
        lines.append("  Keine Warnungen - alle Karten haben eine gute Bild-Zuordnung.")
    if no_match:
        lines.append(f"\n  Karten OHNE Bild-Match ({len(no_match)}):")
        lines.extend(no_match)
    if warnings:
        lines.append(f"\n  Karten mit schwacher Uebereinstimmung (< {WARN_THRESHOLD:.0f}%) ({len(warnings)}):")
        lines.extend(warnings)

    lines.append("")
    lines.append("-" * 100)
    lines.append(f"NICHT ZUGEORDNETE BILDER ({len(unused_images)})")
    lines.append("-" * 100)
    if not unused_images:
        lines.append("  Alle gefundenen Bilder wurden einer Karte zugeordnet.")
    else:
        for img in unused_images:
            lines.append(f"  - {img.path.relative_to(BASE_DIR).as_posix()}")

    lines.append("")
    lines.append("=" * 100)
    lines.append("Dies ist ein DRY RUN. Es wurden KEINE Dateien kopiert oder veraendert.")
    lines.append("Wenn die Zuordnung oben passt, fuehre das Skript erneut mit '--execute' aus.")
    lines.append("=" * 100)

    mapping_file.write_text("\n".join(lines), encoding="utf-8")


# --------------------------------------------------------------------------
# Modus 2: EXECUTE -> Bilder kopieren + JSON schreiben
# --------------------------------------------------------------------------

def slugify(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def try_convert_to_webp(src: Path, target_path: Path) -> bool:
    try:
        from PIL import Image
    except ImportError:
        print(f"  WARNUNG: Pillow nicht installiert - kann {src.name} nicht nach .webp konvertieren.")
        return False
    try:
        with Image.open(src) as im:
            im.save(target_path, "WEBP")
        return True
    except Exception as exc:  # noqa: BLE001
        print(f"  WARNUNG: Konvertierung von {src.name} fehlgeschlagen: {exc}")
        return False


def execute_copy_and_export(cards, pack_ranges):
    CARDS_OUT_DIR.mkdir(parents=True, exist_ok=True)

    copied = 0
    skipped = []
    card_entries = []

    for c in cards:
        target_name = f"card_{c.global_id:03d}.webp"
        target_path = CARDS_OUT_DIR / target_name
        image_rel = None

        if c.assigned_image is not None:
            src = c.assigned_image.path
            if src.suffix.lower() == ".webp":
                shutil.copy2(src, target_path)
                copied += 1
                image_rel = f"cards/{target_name}"
            elif try_convert_to_webp(src, target_path):
                copied += 1
                image_rel = f"cards/{target_name}"
            else:
                skipped.append(c)
        else:
            skipped.append(c)

        card_entries.append(
            {
                "id": f"card_{c.global_id:03d}",
                "pack": c.pack,
                "packId": slugify(c.pack),
                "text": c.text,
                "hex": c.hex_norm,
                "image": image_rel,
                "matched": c.assigned_image is not None,
                "matchScore": round(c.match_score, 1),
                "excelRow": c.row,
            }
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    CARDS_JSON.write_text(json.dumps(card_entries, ensure_ascii=False, indent=2), encoding="utf-8")

    packs_entries = []
    for pr in pack_ranges:
        pack_cards = [c for c in card_entries if c["pack"] == pr["name"]]
        packs_entries.append(
            {
                "id": slugify(pr["name"]),
                "name": pr["name"],
                "cardCount": len(pack_cards),
                "cardIds": [c["id"] for c in pack_cards],
            }
        )
    PACKS_JSON.write_text(json.dumps(packs_entries, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"[EXECUTE] {copied} Bilder kopiert nach {CARDS_OUT_DIR}")
    if skipped:
        print(f"[EXECUTE] {len(skipped)} Karten OHNE Bild (siehe cards.json, 'image': null):")
        for c in skipped:
            print(f"    - card_{c.global_id:03d}: {c.text}")
    print(f"[EXECUTE] Geschrieben: {CARDS_JSON}")
    print(f"[EXECUTE] Geschrieben: {PACKS_JSON}")


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="PlayPointy Karten/Bilder Verarbeitung")
    parser.add_argument(
        "--execute",
        action="store_true",
        help="Fuehrt Modus 2 aus: kopiert Bilder + erstellt cards.json/packs.json. "
        "Ohne dieses Flag laeuft nur der DRY RUN (Modus 1).",
    )
    args = parser.parse_args()

    if not EXCEL_PATH.exists():
        raise SystemExit(f"Excel-Datei nicht gefunden: {EXCEL_PATH}")
    if not PACKS_DIR.exists():
        raise SystemExit(f"Packs-Ordner nicht gefunden: {PACKS_DIR}")

    print("Lese Excel-Datei ein ...")
    cards, pack_ranges = load_all_cards(EXCEL_PATH)
    print(f"  {len(cards)} Karten gefunden ueber {len(pack_ranges)} Packs.")

    print("Durchsuche Packs/-Ordner nach Bildern ...")
    images = collect_images(PACKS_DIR)
    print(f"  {len(images)} Bilddateien gefunden.")

    print("Fuehre Fuzzy-Matching durch (difflib.SequenceMatcher) ...")
    match_cards_to_images(cards, images)
    matched = sum(1 for c in cards if c.assigned_image is not None)
    print(f"  {matched}/{len(cards)} Karten erfolgreich zugeordnet.")

    write_mapping_check(cards, images, pack_ranges, MAPPING_FILE)

    if not args.execute:
        print(f"\nDRY RUN abgeschlossen. Bitte pruefe: {MAPPING_FILE}")
        print("Wenn alles passt, starte das Skript erneut mit --execute.")
    else:
        execute_copy_and_export(cards, pack_ranges)
        print("\nEXECUTE abgeschlossen.")


if __name__ == "__main__":
    main()
